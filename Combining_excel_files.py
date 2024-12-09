from tkinter import *
from tkinter import messagebox, filedialog
import os
from openpyxl import Workbook, open, utils, load_workbook, worksheet
from sys import exit

#The app for combining excel files. \
#With thanks to my best friend DD

root = Tk()
root['bg'] = '#fafafa'
root.title('Объединение эксель файлов с расширением .xlsx')
root.wm_attributes('-alpha', 1.0)
root.geometry('400x400')

frame_1 = Frame(root)
frame_1.pack(side=TOP, fill=X)

frame_2 = Frame(root)
frame_2.pack(side=TOP, fill=X)


def choose_directory():
    """This function cheks the files and create the app window"""
    global dr
    count = 0
    while count == 0:
        dr = filedialog.askdirectory()
        try:
            for file_name in os.listdir(dr):
                if file_name[-4:] == 'xlsx':
                    wb_work = open((os.path.join(dr, f'{file_name}')), read_only=True)
                    ws_work = wb_work.active
                    columns = ws_work.max_column
                    count += 1
                    break

            if count == 0:
                raise Exception('no excel files')

        except Exception:
            error_dr = messagebox.askretrycancel(title='Ошибка!',
                                      message='''В данной папке не найдены эксель-файлы.\
                                              Повторить выбор нужной папки?\n
                                              Для выхода нажмите - "отмена".''')

            if error_dr == True:
                continue
            else:
                exit()

    column_label = Label(frame_1, text='Выберите номера необходимых к объединению столбцов: ')
    column_label.pack(side=TOP, anchor=CENTER)

    r = 0
    c = 0
    global variables
    variables = []
    checkbuttons = []

    for col in range(1, columns+1):
        var = IntVar()
        variables.append(var)
        cb = Checkbutton(frame_2, text=utils.get_column_letter(col), variable=var)
        checkbuttons.append(cb)
        cb.select()
        cb.grid(row=r, column=c)
        r += 1
        if r == 6:
            c += 1
            r = 0

    global file_name_input
    file_name_input = Entry(frame_2, bg='white', width=50)
    file_name_input.insert(0, "Final_file")
    file_name_final = Label(frame_2, text='Введите название итогового файла: ')
    file_name_final.grid(row=7, column=0)
    file_name_input.grid(row=8, column=0)

    bn_start = Button(frame_2, text='Запуск', bg='green', command=start)
    bn_start.grid(row=9, column=0)


wb_final = Workbook()
ws_final = wb_final.active

check = False
def start():
    """This function reads and adds information in excel"""
    col_numb = [i[0] for i in enumerate(variables) if i[1].get() != 0]
    try:
        for file_name in os.listdir(dr):
            if file_name[-4:] == 'xlsx':
                wb_work = load_workbook((os.path.join(dr, f'{file_name}')), data_only=True)
                ws_work = wb_work.active

                final_info = []
                final_info.append([file_name])

                for row in ws_work.iter_rows():
                    info = [cell.value for (col, cell) in list(enumerate(row)) if (
                                        col in col_numb)]
                    final_info.append(info)

                for data in final_info:
                    ws_final.append(data)

        wb_final.save(os.path.join(dr, f'{file_name_input.get()}.xlsx'))
        check = True

        if check == True:
            sucсess_mess = messagebox.showinfo(title='Завершено!',
                                               message='Обобщение данных успешно завершено.')
        else:
            raise Exception()
    except Exception:
        error_mes = messagebox.showerror(title='Ошибка!',
                                             message='Ошибка записи данных.')


#create button to choose directory
bn_dir = Button(frame_1, text='Выбор директории', bg='green', command=choose_directory)
bn_dir.pack(anchor=CENTER, fill=X)


if __name__ == '__main__':
    root.mainloop()